"""
Export Functionality
====================
Export analysis results to PDF and Excel formats.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import io
import json

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_to_pdf(
    analysis_results: Dict[str, Any],
    run_info: Dict[str, Any],
    output_path: Optional[Path] = None
) -> bytes:
    """
    Export analysis results to PDF.
    
    Args:
        analysis_results: Dictionary with all analysis results
        run_info: Run information dictionary
        output_path: Optional path to save PDF (if None, returns bytes)
        
    Returns:
        PDF bytes if output_path is None, otherwise None
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer if output_path is None else str(output_path), pagesize=letter)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e1e2e'),
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph("Comprehensive Analysis Report", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Run Info
    run_info_text = f"""
    <b>Run ID:</b> {run_info.get('run_id', 'N/A')}<br/>
    <b>Name:</b> {run_info.get('name', 'N/A')}<br/>
    <b>Date:</b> {run_info.get('created_at', 'N/A')[:19] if run_info.get('created_at') else 'N/A'}<br/>
    <b>Watchlist:</b> {run_info.get('watchlist', 'N/A')}<br/>
    """
    story.append(Paragraph(run_info_text, styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Performance Attribution
    if 'attribution' in analysis_results:
        story.append(Paragraph("Performance Attribution", styles['Heading2']))
        attr = analysis_results['attribution']
        if isinstance(attr, dict) and 'results' in attr:
            attr_data = attr['results']
            data = [
                ['Component', 'Contribution (%)'],
                ['Total Return', f"{attr_data.get('total_return', 0)*100:.2f}"],
                ['Factor Attribution', f"{attr_data.get('factor_attribution', 0)*100:.2f}"],
                ['Sector Attribution', f"{attr_data.get('sector_attribution', 0)*100:.2f}"],
                ['Stock Selection', f"{attr_data.get('stock_selection_attribution', 0)*100:.2f}"],
                ['Timing', f"{attr_data.get('timing_attribution', 0)*100:.2f}"],
            ]
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2*inch))
    
    # Benchmark Comparison
    if 'benchmark_comparison' in analysis_results:
        story.append(PageBreak())
        story.append(Paragraph("Benchmark Comparison", styles['Heading2']))
        bench = analysis_results['benchmark_comparison']
        if isinstance(bench, dict) and 'results' in bench:
            bench_data = bench['results']
            data = [
                ['Metric', 'Portfolio', 'Benchmark', 'Difference'],
                ['Return (%)', 
                 f"{bench_data.get('portfolio_return', 0)*100:.2f}",
                 f"{bench_data.get('benchmark_return', 0)*100:.2f}",
                 f"{bench_data.get('alpha', 0)*100:.2f}"],
                ['Volatility (%)',
                 f"{bench_data.get('portfolio_volatility', 0)*100:.2f}",
                 f"{bench_data.get('benchmark_volatility', 0)*100:.2f}",
                 f"{(bench_data.get('portfolio_volatility', 0) - bench_data.get('benchmark_volatility', 0))*100:.2f}"],
                ['Sharpe Ratio',
                 f"{bench_data.get('portfolio_sharpe', 0):.2f}",
                 f"{bench_data.get('benchmark_sharpe', 0):.2f}",
                 f"{bench_data.get('portfolio_sharpe', 0) - bench_data.get('benchmark_sharpe', 0):.2f}"],
                ['Beta',
                 f"{bench_data.get('beta', 0):.2f}",
                 "1.00",
                 f"{bench_data.get('beta', 0) - 1.0:.2f}"],
            ]
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2*inch))
    
    # Factor Exposure
    if 'factor_exposure' in analysis_results:
        story.append(PageBreak())
        story.append(Paragraph("Factor Exposure", styles['Heading2']))
        factor = analysis_results['factor_exposure']
        if isinstance(factor, dict) and 'results' in factor:
            factor_data = factor['results']
            exposures = factor_data.get('exposures', {})
            data = [['Factor', 'Exposure', 'Contribution to Return (%)', 'Contribution to Risk (%)']]
            for factor_name, exposure in exposures.items():
                data.append([
                    factor_name,
                    f"{exposure.get('exposure', 0):.3f}",
                    f"{exposure.get('contribution_to_return', 0)*100:.2f}",
                    f"{exposure.get('contribution_to_risk', 0)*100:.2f}"
                ])
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2*inch))
    
    # Build PDF
    doc.build(story)
    
    if output_path is None:
        buffer.seek(0)
        return buffer.read()
    return None


def export_to_excel(
    analysis_results: Dict[str, Any],
    run_info: Dict[str, Any],
    output_path: Optional[Path] = None
) -> bytes:
    """
    Export analysis results to Excel with formatting.
    
    Args:
        analysis_results: Dictionary with all analysis results
        run_info: Run information dictionary
        output_path: Optional path to save Excel (if None, returns bytes)
        
    Returns:
        Excel bytes if output_path is None, otherwise None
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Comprehensive Analysis Report"])
    ws_summary.append([])
    ws_summary.append(["Run ID", run_info.get('run_id', 'N/A')])
    ws_summary.append(["Name", run_info.get('name', 'N/A')])
    ws_summary.append(["Date", run_info.get('created_at', 'N/A')[:19] if run_info.get('created_at') else 'N/A'])
    ws_summary.append(["Watchlist", run_info.get('watchlist', 'N/A')])
    
    # Performance Attribution Sheet
    if 'attribution' in analysis_results:
        ws_attr = wb.create_sheet("Performance Attribution")
        attr = analysis_results['attribution']
        if isinstance(attr, dict) and 'results' in attr:
            attr_data = attr['results']
            ws_attr.append(["Component", "Contribution (%)"])
            ws_attr.append(["Total Return", attr_data.get('total_return', 0)*100])
            ws_attr.append(["Factor Attribution", attr_data.get('factor_attribution', 0)*100])
            ws_attr.append(["Sector Attribution", attr_data.get('sector_attribution', 0)*100])
            ws_attr.append(["Stock Selection", attr_data.get('stock_selection_attribution', 0)*100])
            ws_attr.append(["Timing", attr_data.get('timing_attribution', 0)*100])
            
            # Format header
            for cell in ws_attr[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
    
    # Benchmark Comparison Sheet
    if 'benchmark_comparison' in analysis_results:
        ws_bench = wb.create_sheet("Benchmark Comparison")
        bench = analysis_results['benchmark_comparison']
        if isinstance(bench, dict) and 'results' in bench:
            bench_data = bench['results']
            ws_bench.append(["Metric", "Portfolio", "Benchmark", "Difference"])
            ws_bench.append(["Return (%)", bench_data.get('portfolio_return', 0)*100,
                            bench_data.get('benchmark_return', 0)*100, bench_data.get('alpha', 0)*100])
            ws_bench.append(["Volatility (%)", bench_data.get('portfolio_volatility', 0)*100,
                            bench_data.get('benchmark_volatility', 0)*100,
                            (bench_data.get('portfolio_volatility', 0) - bench_data.get('benchmark_volatility', 0))*100])
            ws_bench.append(["Sharpe Ratio", bench_data.get('portfolio_sharpe', 0),
                            bench_data.get('benchmark_sharpe', 0),
                            bench_data.get('portfolio_sharpe', 0) - bench_data.get('benchmark_sharpe', 0)])
            ws_bench.append(["Beta", bench_data.get('beta', 0), 1.0, bench_data.get('beta', 0) - 1.0])
            
            # Format header
            for cell in ws_bench[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
    
    # Factor Exposure Sheet
    if 'factor_exposure' in analysis_results:
        ws_factor = wb.create_sheet("Factor Exposure")
        factor = analysis_results['factor_exposure']
        if isinstance(factor, dict) and 'results' in factor:
            factor_data = factor['results']
            exposures = factor_data.get('exposures', {})
            ws_factor.append(["Factor", "Exposure", "Contribution to Return (%)", "Contribution to Risk (%)"])
            for factor_name, exposure in exposures.items():
                ws_factor.append([
                    factor_name,
                    exposure.get('exposure', 0),
                    exposure.get('contribution_to_return', 0)*100,
                    exposure.get('contribution_to_risk', 0)*100
                ])
            
            # Format header
            for cell in ws_factor[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
    
    # Rebalancing Sheet
    if 'rebalancing' in analysis_results:
        ws_rebal = wb.create_sheet("Rebalancing")
        rebal = analysis_results['rebalancing']
        if isinstance(rebal, dict) and 'results' in rebal:
            rebal_data = rebal['results']
            ws_rebal.append(["Metric", "Value"])
            ws_rebal.append(["Current Drift (%)", rebal_data.get('current_drift', 0)*100])
            ws_rebal.append(["Average Turnover (%)", rebal_data.get('avg_turnover', 0)*100])
            ws_rebal.append(["Total Transaction Costs (%)", rebal_data.get('total_transaction_costs', 0)*100])
            ws_rebal.append(["Number of Rebalancing Events", rebal_data.get('num_rebalancing_events', 0)])
            
            # Format header
            for cell in ws_rebal[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
    
    # Style Analysis Sheet
    if 'style' in analysis_results:
        ws_style = wb.create_sheet("Style Analysis")
        style = analysis_results['style']
        if isinstance(style, dict) and 'results' in style:
            style_data = style['results']
            ws_style.append(["Metric", "Value"])
            ws_style.append(["Growth/Value Classification", style_data.get('growth_value_classification', 'N/A')])
            ws_style.append(["Size Classification", style_data.get('size_classification', 'N/A')])
            ws_style.append(["Portfolio PE Ratio", style_data.get('portfolio_pe', 0)])
            ws_style.append(["Market Average PE", style_data.get('market_pe', 0)])
            
            # Format header
            for cell in ws_style[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    if output_path is None:
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
    else:
        wb.save(str(output_path))
        return None


def export_to_csv(
    data: pd.DataFrame,
    output_path: Optional[Path] = None
) -> bytes:
    """
    Export DataFrame to CSV.
    
    Args:
        data: DataFrame to export
        output_path: Optional path to save CSV (if None, returns bytes)
        
    Returns:
        CSV bytes if output_path is None, otherwise None
    """
    if output_path is None:
        csv_str = data.to_csv(index=False)
        return csv_str.encode('utf-8')
    else:
        data.to_csv(str(output_path), index=False)
        return None


def export_to_json(
    data: Any,
    output_path: Optional[Path] = None,
    indent: int = 2
) -> bytes:
    """
    Export data to JSON.
    
    Args:
        data: Data to export (dict, list, or DataFrame)
        output_path: Optional path to save JSON (if None, returns bytes)
        indent: JSON indentation (default: 2)
        
    Returns:
        JSON bytes if output_path is None, otherwise None
    """
    # Convert DataFrame to dict if needed
    if isinstance(data, pd.DataFrame):
        json_data = data.to_dict('records')
    else:
        json_data = data
    
    json_str = json.dumps(json_data, indent=indent, default=str)
    
    if output_path is None:
        return json_str.encode('utf-8')
    else:
        with open(str(output_path), 'w') as f:
            f.write(json_str)
        return None

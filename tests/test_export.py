"""
Test Export Functionality
==========================
Test cases for PDF and Excel export.
"""

import pytest
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
import tempfile
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.app.dashboard.export import export_to_pdf, export_to_excel


class TestPDFExport:
    """Test PDF export functionality."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.analysis_results = {
            'attribution': {
                'results': {
                    'total_return': 0.15,
                    'factor_attribution': 0.05,
                    'sector_attribution': 0.04,
                    'stock_selection_attribution': 0.03,
                    'timing_attribution': 0.03
                }
            },
            'benchmark_comparison': {
                'results': {
                    'portfolio_return': 0.15,
                    'benchmark_return': 0.12,
                    'alpha': 0.03,
                    'portfolio_volatility': 0.18,
                    'benchmark_volatility': 0.15,
                    'portfolio_sharpe': 0.83,
                    'benchmark_sharpe': 0.80,
                    'beta': 1.05
                }
            },
            'factor_exposure': {
                'results': {
                    'exposures': {
                        'Market': {'exposure': 0.95, 'contribution_to_return': 0.10, 'contribution_to_risk': 0.12},
                        'Value': {'exposure': 0.25, 'contribution_to_return': 0.02, 'contribution_to_risk': 0.03}
                    }
                }
            },
            'rebalancing': {
                'results': {
                    'current_drift': 0.05,
                    'avg_turnover': 0.20,
                    'total_transaction_costs': 0.002,
                    'num_rebalancing_events': 12
                }
            },
            'style': {
                'results': {
                    'growth_value_classification': 'Value',
                    'size_classification': 'Large Cap',
                    'portfolio_pe': 18.5,
                    'market_pe': 20.0
                }
            }
        }
        
        self.run_info = {
            'run_id': 'test_run_123',
            'name': 'Test Run',
            'created_at': datetime.now().isoformat(),
            'watchlist': 'test_watchlist'
        }
    
    @pytest.mark.skipif(
        not pytest.importorskip('reportlab', reason='reportlab not installed'),
        reason='reportlab not available'
    )
    def test_export_to_pdf_bytes(self):
        """Test PDF export to bytes."""
        pdf_bytes = export_to_pdf(self.analysis_results, self.run_info)
        
        assert pdf_bytes is not None
        assert len(pdf_bytes) > 0
        assert pdf_bytes.startswith(b'%PDF')  # PDF file signature
    
    @pytest.mark.skipif(
        not pytest.importorskip('reportlab', reason='reportlab not installed'),
        reason='reportlab not available'
    )
    def test_export_to_pdf_file(self):
        """Test PDF export to file."""
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            export_to_pdf(self.analysis_results, self.run_info, output_path=Path(tmp.name))
            
            assert Path(tmp.name).exists()
            assert Path(tmp.name).stat().st_size > 0
            
            # Clean up
            Path(tmp.name).unlink()
    
    def test_export_to_pdf_missing_reportlab(self):
        """Test PDF export error when reportlab not available."""
        # This test would need to mock the import
        pass


class TestExcelExport:
    """Test Excel export functionality."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.analysis_results = {
            'attribution': {
                'results': {
                    'total_return': 0.15,
                    'factor_attribution': 0.05,
                    'sector_attribution': 0.04,
                    'stock_selection_attribution': 0.03,
                    'timing_attribution': 0.03
                }
            },
            'benchmark_comparison': {
                'results': {
                    'portfolio_return': 0.15,
                    'benchmark_return': 0.12,
                    'alpha': 0.03,
                    'portfolio_volatility': 0.18,
                    'benchmark_volatility': 0.15,
                    'portfolio_sharpe': 0.83,
                    'benchmark_sharpe': 0.80,
                    'beta': 1.05
                }
            },
            'factor_exposure': {
                'results': {
                    'exposures': {
                        'Market': {'exposure': 0.95, 'contribution_to_return': 0.10, 'contribution_to_risk': 0.12},
                        'Value': {'exposure': 0.25, 'contribution_to_return': 0.02, 'contribution_to_risk': 0.03}
                    }
                }
            },
            'rebalancing': {
                'results': {
                    'current_drift': 0.05,
                    'avg_turnover': 0.20,
                    'total_transaction_costs': 0.002,
                    'num_rebalancing_events': 12
                }
            },
            'style': {
                'results': {
                    'growth_value_classification': 'Value',
                    'size_classification': 'Large Cap',
                    'portfolio_pe': 18.5,
                    'market_pe': 20.0
                }
            }
        }
        
        self.run_info = {
            'run_id': 'test_run_123',
            'name': 'Test Run',
            'created_at': datetime.now().isoformat(),
            'watchlist': 'test_watchlist'
        }
    
    @pytest.mark.skipif(
        not pytest.importorskip('openpyxl', reason='openpyxl not installed'),
        reason='openpyxl not available'
    )
    def test_export_to_excel_bytes(self):
        """Test Excel export to bytes."""
        excel_bytes = export_to_excel(self.analysis_results, self.run_info)
        
        assert excel_bytes is not None
        assert len(excel_bytes) > 0
        # Excel files start with PK (ZIP signature)
        assert excel_bytes.startswith(b'PK')
    
    @pytest.mark.skipif(
        not pytest.importorskip('openpyxl', reason='openpyxl not installed'),
        reason='openpyxl not available'
    )
    def test_export_to_excel_file(self):
        """Test Excel export to file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            export_to_excel(self.analysis_results, self.run_info, output_path=Path(tmp.name))
            
            assert Path(tmp.name).exists()
            assert Path(tmp.name).stat().st_size > 0
            
            # Verify it's a valid Excel file
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            assert len(wb.sheetnames) > 0
            assert 'Summary' in wb.sheetnames
            
            # Clean up
            Path(tmp.name).unlink()
    
    @pytest.mark.skipif(
        not pytest.importorskip('openpyxl', reason='openpyxl not installed'),
        reason='openpyxl not available'
    )
    def test_excel_multiple_sheets(self):
        """Test that Excel export creates multiple sheets."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            export_to_excel(self.analysis_results, self.run_info, output_path=Path(tmp.name))
            
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            
            # Should have multiple sheets
            assert len(wb.sheetnames) >= 3
            assert 'Summary' in wb.sheetnames
            assert 'Performance Attribution' in wb.sheetnames or 'Attribution' in wb.sheetnames
            
            # Clean up
            Path(tmp.name).unlink()
